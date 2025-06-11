# integrations/sharepoint/client.py
"""
Office365-REST-Python-Client based SharePoint integration using NATIVE SDK SCHEMAS.

This module leverages the rich File and Folder objects provided by the Office365 SDK
directly, rather than creating custom wrapper classes. The SDK's native objects have
all the properties we need!

Native SDK objects we use:
- File: Has unique_id, name, serverRelativeUrl, length, time_created, time_last_modified,
  author, parent_folder, list_id, site_id, web_id, major_version, minor_version, etc.
- Folder: Has unique_id, name, serverRelativeUrl, time_created, time_last_modified,
  files, folders, parent_folder, etc.
- FileCollection & FolderCollection: For iterating over multiple items
"""

import os
from pathlib import Path
from typing import List, Optional, Tuple

from office365.runtime.auth.client_credential import ClientCredential
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
from office365.sharepoint.folders.folder import Folder


class Office365SharePointClient:
    """
    Modern SharePoint client using Office365-REST-Python-Client SDK with NATIVE SCHEMAS.

    This client returns the SDK's native File and Folder objects directly, which have
    all the rich properties and methods we need. No custom wrapper classes required!
    """

    def __init__(
        self,
        site_url: str = "https://jgiquality.sharepoint.com/sites/JGI",
        auth_method: str = "app_only",
    ):
        """
        Initialize SharePoint client with Office365 SDK.

        Args:
            site_url: Full SharePoint site URL
            auth_method: "app_only" or "user_credential"
        """
        self.site_url = site_url
        self.auth_method = auth_method
        self._context: Optional[ClientContext] = None

    def _get_context(self) -> ClientContext:
        """Get authenticated ClientContext."""
        if self._context is None:
            if self.auth_method == "app_only":
                self._context = self._get_app_context()
            else:
                raise NotImplementedError("User credential auth not implemented yet")

        return self._context

    def _get_app_context(self) -> ClientContext:
        """Get context using app-only authentication."""
        client_id = os.getenv("AZURE_CLIENT_ID")
        client_secret = os.getenv("AZURE_CLIENT_SECRET")

        if not all([client_id, client_secret]):
            raise ValueError(
                "Missing Azure app credentials: AZURE_CLIENT_ID and AZURE_CLIENT_SECRET required"
            )

        credentials = ClientCredential(client_id, client_secret)
        context = ClientContext(self.site_url).with_credentials(credentials)
        return context

    def _execute_query_with_error_handling(self, context: ClientContext) -> None:
        """Execute query with proper error handling."""
        try:
            context.execute_query()
        except Exception as e:
            error_msg = str(e)
            if "404" in error_msg or "Not Found" in error_msg:
                raise FileNotFoundError(f"SharePoint resource not found: {e}")
            elif "403" in error_msg or "Forbidden" in error_msg:
                raise PermissionError(f"Access denied to SharePoint resource: {e}")
            else:
                raise RuntimeError(f"SharePoint operation failed: {e}")

    def get_file_by_url(self, relative_url: str) -> File:
        """
        Get file object by server-relative URL.

        Args:
            relative_url: Server-relative URL to the file

        Returns:
            Native Office365 File object with all rich properties:
            - file.unique_id, file.name, file.serverRelativeUrl
            - file.length, file.time_created, file.time_last_modified
            - file.author, file.parent_folder, file.major_version, etc.

        Raises:
            FileNotFoundError: If file doesn't exist
        """
        context = self._get_context()

        # Use Office365 SDK to get file - returns native File object
        file_obj = context.web.get_file_by_server_relative_url(relative_url)
        context.load(file_obj)
        self._execute_query_with_error_handling(context)

        if not file_obj.exists:
            raise FileNotFoundError(f"File not found: {relative_url}")

        return file_obj  # Return native SDK File object!

    def search_files(
        self, query: str, folder_path: Optional[Path] = None
    ) -> List[File]:
        """
        Search for files using SharePoint folder traversal.

        Args:
            query: Search query string (filename contains)
            folder_path: Optional folder to restrict search

        Returns:
            List of native Office365 File objects with rich properties
        """
        context = self._get_context()

        if folder_path:
            # Search within specific folder
            folder = context.web.get_folder_by_server_relative_url(folder_path)
            context.load(folder)
            context.load(folder.files)
            self._execute_query_with_error_handling(context)
            files = folder.files
        else:
            # Search in document library root
            list_obj = context.web.lists.get_by_title("Documents")
            context.load(list_obj.root_folder.files)
            self._execute_query_with_error_handling(context)
            files = list_obj.root_folder.files

        results = []
        for file_obj in files:
            if query.lower() in file_obj.name.lower():
                context.load(file_obj)  # Load full file properties
                self._execute_query_with_error_handling(context)
                results.append(file_obj)  # Native File object!

        return results

    def get_folder_contents(self, folder_path: Path) -> Tuple[List[File], List[Folder]]:
        """
        Get contents of a folder.

        Args:
            folder_path (Path): Server-relative path to folder

        Returns:
            Tuple of (files, subfolders) as native SDK objects:
            - List[File]: Native File objects with all properties
            - List[Folder]: Native Folder objects with all properties
        """
        context = self._get_context()

        folder = context.web.get_folder_by_server_relative_url(folder_path)
        context.load(folder)
        context.load(folder.files)
        context.load(folder.folders)
        self._execute_query_with_error_handling(context)

        files = []
        for file_obj in folder.files:
            context.load(file_obj)  # Load full file properties
            self._execute_query_with_error_handling(context)
            files.append(file_obj)  # Native File object!

        folders = []
        for folder_obj in folder.folders:
            context.load(folder_obj)  # Load full folder properties
            self._execute_query_with_error_handling(context)
            folders.append(folder_obj)  # Native Folder object!

        return files, folders

    def download_file_content(self, relative_url: str) -> bytes:
        """
        Download file content as bytes.

        Args:
            relative_url: Server-relative URL to file

        Returns:
            File content as bytes
        """
        context = self._get_context()

        file_obj = context.web.get_file_by_server_relative_url(relative_url)
        content_result = file_obj.get_content()
        self._execute_query_with_error_handling(context)

        return content_result.value

    def download_file_to_path(self, relative_url: str, local_path: Path) -> None:
        """
        Download file to local filesystem.

        Args:
            relative_url (str): Server-relative URL to file
            local_path (Path): Local filesystem path to save file
        """
        content = self.download_file_content(relative_url)

        with open(local_path, "wb") as local_file:
            local_file.write(content)

    def upload_file(
        self, local_path: Path, target_folder: str, filename: Optional[str] = None
    ) -> File:
        """
        Upload file to SharePoint.

        Args:
            local_path: Path to local file
            target_folder: Target SharePoint folder path
            filename: Optional custom filename (uses local filename if not provided)

        Returns:
            Native Office365 File object of uploaded file
        """
        context = self._get_context()

        if filename is None:
            filename = os.path.basename(local_path)

        target_folder_obj = context.web.get_folder_by_server_relative_url(target_folder)

        with open(local_path, "rb") as file_content:
            uploaded_file = target_folder_obj.files.upload(file_content, filename)
            self._execute_query_with_error_handling(context)

        context.load(uploaded_file)  # Load full file properties
        self._execute_query_with_error_handling(context)

        return uploaded_file  # Native File object!


# Factory functions for backward compatibility with existing API
def create_sharepoint_client(
    site_url: Optional[str] = None,
) -> Office365SharePointClient:
    """Create SharePoint client with default site URL."""
    if site_url is None:
        site_url = os.getenv(
            "SHAREPOINT_SITE_URL", "https://jgiquality.sharepoint.com/sites/JGI"
        )

    return Office365SharePointClient(site_url)


def get_pyro_standards_files() -> List[File]:
    """
    Get all files from Pyro Standards folder using Office365 SDK.

    Returns:
        List of native Office365 File objects with rich properties:
        - file.unique_id, file.name, file.serverRelativeUrl
        - file.length, file.time_created, file.time_last_modified
        - file.author, file.parent_folder, file.major_version, etc.
    """
    client = create_sharepoint_client()

    # Get Pyro Standards folder contents
    pyro_standards_path = Path("/sites/JGI/Shared Documents/Pyro/Pyro_Standards")
    files, _ = client.get_folder_contents(pyro_standards_path)

    return files


def get_wiresetcerts_file() -> File:
    """
    Get WireSetCerts.xlsx file using Office365 SDK.

    Returns:
        Native Office365 File object with all rich properties
    """
    client = create_sharepoint_client()

    wiresetcerts_path = "/sites/JGI/Shared Documents/Pyro/WireSetCerts.xlsx"
    return client.get_file_by_url(wiresetcerts_path)


def download_wiresetcerts_content() -> bytes:
    """
    Download WireSetCerts.xlsx content as bytes.

    Returns:
        File content as bytes
    """
    client = create_sharepoint_client()

    wiresetcerts_path = "/sites/JGI/Shared Documents/Pyro/WireSetCerts.xlsx"
    return client.download_file_content(wiresetcerts_path)


# Utility functions to extract common properties from native SDK objects
def extract_file_info_dict(file_obj: File) -> dict:
    """
    Extract file information from native File object into a dictionary.

    This is useful for backward compatibility with code that expects dictionaries.

    Args:
        file_obj: Native Office365 File object

    Returns:
        Dictionary with file metadata
    """
    return {
        "id": file_obj.unique_id or "",
        "name": file_obj.name or "",
        "serverRelativeUrl": file_obj.serverRelativeUrl or "",
        "size": file_obj.length or 0,
        "lastModifiedDateTime": (
            file_obj.time_last_modified.isoformat()
            if file_obj.time_last_modified
            else ""
        ),
        "createdDateTime": (
            file_obj.time_created.isoformat() if file_obj.time_created else ""
        ),
        "webUrl": (
            f"https://jgiquality.sharepoint.com{file_obj.serverRelativeUrl}"
            if file_obj.serverRelativeUrl
            else ""
        ),
        "downloadUrl": (
            f"https://jgiquality.sharepoint.com{file_obj.serverRelativeUrl}"
            if file_obj.serverRelativeUrl
            else ""
        ),
        "author": getattr(file_obj.author, "title", "") if file_obj.author else "",
        "majorVersion": file_obj.major_version,
        "minorVersion": file_obj.minor_version,
        "listId": file_obj.list_id,
        "siteId": file_obj.site_id,
        "webId": file_obj.web_id,
    }


def extract_folder_info_dict(folder_obj: Folder) -> dict:
    """
    Extract folder information from native Folder object into a dictionary.

    Args:
        folder_obj: Native Office365 Folder object

    Returns:
        Dictionary with folder metadata
    """
    return {
        "id": folder_obj.unique_id or "",
        "name": folder_obj.name or "",
        "serverRelativeUrl": folder_obj.serverRelativeUrl or "",
        "lastModifiedDateTime": (
            folder_obj.time_last_modified.isoformat()
            if folder_obj.time_last_modified
            else ""
        ),
        "createdDateTime": (
            folder_obj.time_created.isoformat() if folder_obj.time_created else ""
        ),
        "webUrl": (
            f"https://jgiquality.sharepoint.com{folder_obj.serverRelativeUrl}"
            if folder_obj.serverRelativeUrl
            else ""
        ),
        "itemCount": getattr(folder_obj, "item_count", None),
    }


# High-level convenience functions
def get_pyro_standards_excel_file(file_name: str) -> dict:
    """
    Get Excel file content from Pyro_Standards folder with parsing.

    Args:
        file_name: Name of the Excel file to retrieve

    Returns:
        Dictionary with file info and parsed Excel content
    """
    try:
        import io

        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel file parsing")

    # Get all files in Pyro Standards folder
    pyro_files = get_pyro_standards_files()

    # Find the matching file (case insensitive)
    matching_file = None
    for file_obj in pyro_files:
        if file_obj.name and file_obj.name.lower() == file_name.lower():
            matching_file = file_obj
            break

    if not matching_file:
        raise ValueError(f"{file_name} file not found in SharePoint")

    if not matching_file.serverRelativeUrl:
        raise ValueError(f"{file_name} does not have a valid server-relative URL")

    # Download and parse the Excel file
    client = create_sharepoint_client()
    content = client.download_file_content(matching_file.serverRelativeUrl)

    # Create file info dictionary
    file_info = extract_file_info_dict(matching_file)

    # Parse Excel content
    file_like = io.BytesIO(content)
    workbook = openpyxl.load_workbook(file_like, read_only=True)

    sheets = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get sheet data
        headers = []
        sample_data = []

        try:
            all_rows = list(sheet.iter_rows(values_only=True))
            if all_rows:
                headers = list(all_rows[0]) if all_rows[0] else []
                sample_data = [list(row) for row in all_rows[1:6]]
        except (IndexError, TypeError):
            pass

        sheets[sheet_name] = {
            "headers": headers,
            "sample_data": sample_data,
            "total_rows": sheet.max_row or 0,
            "total_columns": sheet.max_column or 0,
        }

    workbook.close()
    return {
        "total_sheets": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames,
        "sheets": sheets,
        "file_info": file_info,
    }


def list_pyro_standards_excel_files() -> List[File]:
    """
    List Excel files in Pyro Standards folder.

    Returns:
        List of File objects
    """
    pyro_files = get_pyro_standards_files()

    excel_files = []
    for file_obj in pyro_files:
        if file_obj.name and file_obj.name.endswith((".xlsx", ".xls", ".xlsm")):
            excel_files.append(file_obj)

    return excel_files


def get_wiresetcerts_content() -> dict:
    """
    Get WireSetCerts.xlsx content with Excel parsing.

    Returns:
        Dictionary with parsed Excel content
    """
    try:
        import io

        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel file parsing")

    # Download content
    content = download_wiresetcerts_content()

    # Parse Excel content
    file_like = io.BytesIO(content)
    workbook = openpyxl.load_workbook(file_like, read_only=True)

    sheets = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get sheet data
        all_rows = list(sheet.iter_rows(values_only=True))
        headers = all_rows[0] if all_rows else []
        sample_data = all_rows[1:6] if len(all_rows) > 1 else []

        sheets[sheet_name] = {
            "total_rows": len(all_rows),
            "total_columns": len(headers) if headers else 0,
            "headers": list(headers),
            "sample_data": [list(row) for row in sample_data],
        }

    workbook.close()
    return {
        "total_sheets": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames,
        "sheets": sheets,
    }


def get_wiresetcerts_file_reference() -> dict:
    """
    Get WireSetCerts.xlsx file reference.

    Returns:
        Dictionary with file metadata
    """
    wiresetcerts_file = get_wiresetcerts_file()
    return extract_file_info_dict(wiresetcerts_file)


def get_pyro_file_reference(file_path: str) -> dict:
    """
    Get file reference from Pyro folder.

    Args:
        file_path: Path to file within Pyro folder

    Returns:
        Dictionary with file metadata
    """
    client = create_sharepoint_client()
    server_relative_url = f"/sites/JGI/Shared Documents/Pyro/{file_path}"
    file_obj = client.get_file_by_url(server_relative_url)
    return extract_file_info_dict(file_obj)


def get_pyro_standards_file_reference(file_path: str) -> dict:
    """
    Get file reference from Pyro Standards folder.

    Args:
        file_path: Path to file within Pyro Standards folder

    Returns:
        Dictionary with file metadata
    """
    client = create_sharepoint_client()
    server_relative_url = f"/sites/JGI/Shared Documents/Pyro/Pyro_Standards/{file_path}"
    file_obj = client.get_file_by_url(server_relative_url)
    return extract_file_info_dict(file_obj)


def search_pyro_files(query: str) -> List[dict]:
    """
    Search for files in Pyro folder.

    Args:
        query: Search query string

    Returns:
        List of dictionaries with file metadata
    """
    client = create_sharepoint_client()
    pyro_folder_path = Path("/sites/JGI/Shared Documents/Pyro")
    files = client.search_files(query, pyro_folder_path)
    return [extract_file_info_dict(f) for f in files]


def list_pyro_folder_contents(folder_path: Path) -> List[dict]:
    """
    List contents of a folder within Pyro.

    Args:
        folder_path (Path): Path to folder within Pyro

    Returns:
        List of dictionaries with file and folder metadata
    """
    client = create_sharepoint_client()
    server_relative_path = Path(f"/sites/JGI/Shared Documents/Pyro/{folder_path}")
    if not server_relative_path.is_absolute():
        raise ValueError("folder_path must be an absolute path within Pyro")
    files, folders = client.get_folder_contents(server_relative_path)

    results = []

    # Add files
    for file_obj in files:
        results.append(extract_file_info_dict(file_obj))

    # Add folders
    for folder_obj in folders:
        folder_dict = extract_folder_info_dict(folder_obj)
        folder_dict["folder"] = {}  # Mark as folder
        results.append(folder_dict)

    return results
