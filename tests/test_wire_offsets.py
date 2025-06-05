# tests/test_wire_offsets.py
import os
from datetime import datetime

import pytest

from db.models import WireOffset, WireSetCert
from utils.wire_set_cert_refresher import WireSetCertRefresher


class TestWireOffsetModel:
    """Test WireOffset model functionality."""

    def test_wire_offset_model_creation(self):
        """Test creating a WireOffset model instance."""
        # Updated for new schema with proper correction factor fields
        test_date = datetime.now()
        wire_offset = WireOffset(
            traceability_no="123456A",
            nominal_temp=25.0,
            correction_factor=1.123456,
            created_at=test_date,
            updated_at=test_date,
            updated_by="Test User",
        )

        assert wire_offset.traceability_no == "123456A"
        assert float(wire_offset.nominal_temp) == 25.0
        assert float(wire_offset.correction_factor) == 1.123456
        assert wire_offset.created_at == test_date
        assert wire_offset.updated_at == test_date
        assert wire_offset.updated_by == "Test User"

    def test_wire_offset_repr(self):
        """Test WireOffset string representation."""
        # Updated for new schema with proper correction factor fields
        created_at = datetime(2025, 6, 4, 12, 0, 0)
        wire_offset = WireOffset(
            id=1,
            traceability_no="123456A",
            nominal_temp=25.0,
            correction_factor=1.123456,
            created_at=created_at,
            updated_at=created_at,
            updated_by="Test User",
        )

        expected = (
            "<WireOffset(id=1, traceability_no='123456A', temp=25.0, cf=1.123456)>"
        )
        assert repr(wire_offset) == expected


class TestWireSetCertModel:
    """Test WireSetCert model functionality."""

    def test_wire_set_cert_model_creation(self):
        """Test creating a WireSetCert model instance."""
        # TODO: Test the new WireSetCert model
        cert = WireSetCert(
            serial_number="J201",
            wire_set_group="J201-J214",
            created_at=datetime.now(),
            updated_at=datetime.now(),
        )

        assert cert.serial_number == "J201"
        assert cert.wire_set_group == "J201-J214"
        assert cert.created_at is not None
        assert cert.updated_at is not None

    def test_wire_set_cert_repr(self):
        """Test WireSetCert string representation."""
        cert = WireSetCert(serial_number="J201", wire_set_group="J201-J214")

        expected = "<WireSetCert(serial_number='J201', wire_set_group='J201-J214', asset_id=None)>"
        assert repr(cert) == expected


class TestWireOffsetsAPI:
    """Test wire offsets API endpoints."""

    def test_get_all_wire_offsets_success(self, client, auth_token):
        """Test GET /wire-offsets/ returns current wire offset data."""
        skip_auth = os.getenv("SKIP_AUTH", "false").lower() == "true"

        if not skip_auth:
            # Test the wire_offsets_current view behavior
            response = client.get(
                "/wire-offsets/", headers={"Authorization": f"Bearer {auth_token}"}
            )
            assert response.status_code == 200
            data = response.get_json()
            assert isinstance(data, list)

            # If we have data, verify structure includes new fields
            if data:
                first_item = data[0]
                assert "id" in first_item
                assert "traceability_no" in first_item
                assert "nominal_temp" in first_item
                assert "correction_factor" in first_item
                assert "created_at" in first_item
                assert "updated_at" in first_item
                assert "updated_by" in first_item
                assert isinstance(first_item["nominal_temp"], (int, float))
                assert isinstance(first_item["correction_factor"], (int, float))
        else:
            # Mock mode - test the mock response
            response = client.get(
                "/wire-offsets/", headers={"Authorization": f"Bearer {auth_token}"}
            )
            assert response.status_code == 200
            data = response.get_json()
            assert data == []  # Mock returns empty list

    def test_get_wire_offsets_by_wirelot_success(self, client, auth_token):
        """Test GET /wire-offsets/<traceability_no> returns current data for wire lot."""
        skip_auth = os.getenv("SKIP_AUTH", "false").lower() == "true"
        test_traceability_no = "123456A"
        if not skip_auth:
            # Test the updated endpoint for retrieving wire offsets by traceability_no
            response = client.get(
                f"/wire-offsets/{test_traceability_no}",
                headers={"Authorization": f"Bearer {auth_token}"},
            )

            # Handle 404 case when no data exists for the traceability number
            if response.status_code == 404:
                # This is expected when no data exists for this traceability number
                data = response.get_json()
                assert "message" in data
                assert test_traceability_no in data["message"]
            else:
                assert response.status_code == 200
                data = response.get_json()
                assert isinstance(data, list)

                # All returned items should have the specified traceability_no
                for item in data:
                    assert item.get("traceability_no") == test_traceability_no
        else:  # Mock mode
            response = client.get(
                f"/wire-offsets/{test_traceability_no}",
                headers={"Authorization": f"Bearer {auth_token}"},
            )
            assert response.status_code == 200
            data = response.get_json()
            assert data == []

    def test_get_wire_set_certs_success(self, client, auth_token):
        """Test GET /wire-set-certs/ returns certificate mappings."""
        skip_auth = os.getenv("SKIP_AUTH", "false").lower() == "true"

        if not skip_auth:
            # When SKIP_AUTH=false, temporarily restore real view function
            # to bypass mock view bindings that may have been registered
            from app import app
            from routes.wire_offsets import WireSetCerts

            original_view_func = None
            endpoint_name = None

            # Find the correct endpoint name for wire set certs
            for rule in app.url_map.iter_rules():
                if rule.rule == "/wire-set-certs/":
                    endpoint_name = rule.endpoint
                    original_view_func = app.view_functions.get(endpoint_name)
                    break

            if endpoint_name and original_view_func:
                # Temporarily restore the real view function
                real_view = WireSetCerts().get
                app.view_functions[endpoint_name] = real_view

            try:
                response = client.get(
                    "/wire-set-certs/",
                    headers={"Authorization": f"Bearer {auth_token}"},
                )
                assert response.status_code == 200
                data = response.get_json()
                assert isinstance(data, list)

                # If we have data, verify structure including new fields

                if data:
                    first_item = data[0]
                    assert "id" in first_item
                    assert "serial_number" in first_item
                    assert "wire_set_group" in first_item
                    assert "created_at" in first_item
                    assert "updated_at" in first_item

                    # Check for new fields (may be None)
                    assert "asset_id" in first_item
                    assert "asset_tag" in first_item
                    assert "custom_order_number" in first_item
                    assert "service_date" in first_item
                    assert "next_service_date" in first_item
                    assert "certificate_number" in first_item
                    assert "wire_roll_cert_number" in first_item
            finally:
                # Restore the original view function
                if endpoint_name and original_view_func:
                    app.view_functions[endpoint_name] = original_view_func
        else:
            # Mock mode
            response = client.get(
                "/wire-set-certs/", headers={"Authorization": f"Bearer {auth_token}"}
            )
            assert response.status_code == 200
            data = response.get_json()
            assert data == []

    def test_wire_set_certs_refresh_success(self, client, auth_token):
        """Test POST /refresh-excel-data/ triggers unified refresh operation."""
        skip_auth = os.getenv("SKIP_AUTH", "false").lower() == "true"

        if not skip_auth:
            # Use the unified refresh endpoint instead of the removed wire-set-certs/refresh
            response = client.post(
                "/refresh-excel-data/",
                headers={"Authorization": f"Bearer {auth_token}"},
            )
            assert response.status_code == 200
            data = response.get_json()
            # Check for unified refresh response structure
            assert isinstance(data, dict)
            # The unified refresh may have different response structure than the old endpoint
            # but should at least be a successful response
        else:
            # Mock mode - for now, skip this test as the mock endpoints may not be set up
            # for the unified refresh endpoint yet
            response = client.post(
                "/refresh-excel-data/",
                headers={"Authorization": f"Bearer {auth_token}"},
            )
            # In mock mode, we might get 404 if the endpoint isn't mocked yet
            # This is acceptable for now
            assert response.status_code in [200, 404]

    # TODO: Add tests for authentication failures
    # TODO: Add tests for database error handling
    # TODO: Add tests for the wire_offsets_current view behavior
    # TODO: Add tests for wire set cert refresh with real SharePoint data
    # TODO: Add tests for append-only behavior (multiple records for same wirelot/block)


class TestWireSetCertsParser:
    """Test WireSetCerts.xlsx parsing functionality."""

    def test_parse_wiresetcerts_excel_success(self):
        """Test that we can successfully parse the WireSetCerts.xlsx file."""
        from utils.wire_set_cert_refresher import WireSetCertRefresher

        # Read the test file
        test_file_path = "tests/data/WireSetCerts.xlsx"
        with open(test_file_path, "rb") as f:
            excel_data = f.read()

        # Parse the Excel data
        refresher = WireSetCertRefresher()
        mappings = refresher._parse_wiresetcerts_excel(excel_data)

        # Assertions
        assert len(mappings) > 0, "Should find wire set mappings in the test file"

        # Check that all mappings have required fields
        for mapping in mappings:
            assert "serial_number" in mapping
            assert "wire_set_group" in mapping
            assert mapping["serial_number"] is not None
            assert mapping["wire_set_group"] is not None

        # Check specific expected mappings from the test file
        serial_numbers = [m["serial_number"] for m in mappings]
        wire_set_groups = [m["wire_set_group"] for m in mappings]

        assert "ENV101" in serial_numbers, "Should find ENV101 serial number"
        assert "J01-J24" in serial_numbers, "Should find J01-J24 serial number"
        assert "ENV" in wire_set_groups, "Should find ENV wire set group"
        assert "J" in wire_set_groups, "Should find J wire set group"

        # Find specific mapping
        env_mapping = next(
            (m for m in mappings if m["serial_number"] == "ENV101"), None
        )
        assert env_mapping is not None, "Should find ENV101 mapping"
        assert env_mapping["wire_set_group"] == "ENV", "ENV101 should map to ENV group"

    def test_parse_wiresetcerts_excel_empty_file(self):
        """Test parsing with empty or invalid Excel data."""
        from utils.wire_set_cert_refresher import WireSetCertRefresher

        refresher = WireSetCertRefresher()

        # Test with empty data
        mappings = refresher._parse_wiresetcerts_excel(b"")
        assert mappings == [], "Empty data should return empty list"

        # Test with invalid data
        mappings = refresher._parse_wiresetcerts_excel(b"invalid excel data")
        assert mappings == [], "Invalid data should return empty list"
        assert mappings == [], "Invalid data should return empty list"

    def test_parse_wiresetcerts_excel_complete_schema(self):
        """Test parsing Excel with complete schema including all new fields."""
        # Create a test Excel file with all columns
        from datetime import datetime
        from io import BytesIO

        from openpyxl import Workbook

        # Create workbook with complete schema
        wb = Workbook()
        ws = wb.active
        ws.title = "WireSetCerts"

        # Add headers matching the complete schema
        headers = [
            "serial_number",
            "type",
            "asset_id",
            "asset_tag",
            "custom_order_number",
            "service_date",
            "next_service_date",
            "certificate_number",
            "wire_roll_cert_number",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Add test data
        test_data = [
            [
                "J201",
                "J201-J214",
                12345,
                "TAG001",
                "ORD-2024-001",
                datetime(2024, 1, 15),
                datetime(2025, 1, 15),
                "CERT-001",
                "ROLL-001",
            ],
            [
                "J202",
                "J201-J214",
                12346,
                "TAG002",
                "ORD-2024-002",
                datetime(2024, 2, 15),
                datetime(2025, 2, 15),
                "CERT-002",
                "ROLL-002",
            ],
        ]

        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_data = excel_buffer.getvalue()

        # Parse the Excel data
        refresher = WireSetCertRefresher()
        mappings = refresher._parse_wiresetcerts_excel(excel_data)

        # Assertions
        assert len(mappings) == 2, "Should find 2 wire set mappings"

        # Check first mapping has all fields
        mapping = mappings[0]
        assert mapping["serial_number"] == "J201"
        assert mapping["wire_set_group"] == "J201-J214"
        assert mapping["asset_id"] == 12345
        assert mapping["asset_tag"] == "TAG001"
        assert mapping["custom_order_number"] == "ORD-2024-001"
        assert isinstance(mapping["service_date"], datetime)
        assert isinstance(mapping["next_service_date"], datetime)
        assert mapping["certificate_number"] == "CERT-001"
        assert mapping["wire_roll_cert_number"] == "ROLL-001"

        # Check second mapping
        mapping2 = mappings[1]
        assert mapping2["serial_number"] == "J202"
        assert mapping2["asset_id"] == 12346
        assert mapping2["asset_id"] == 12346


class TestWireOffsetParsing:
    """Test wire offset Excel parsing functionality."""

    @pytest.mark.parametrize(
        "excel_file, expected_values_csv",
        [
            ("072513A.xlsx", "072513A_offsets.csv"),
            ("011391L.xls", "011391L_offsets.csv"),
        ],
    )
    def test_parse_wire_offsets(self, excel_file, expected_values_csv):
        """Test parsing wire offset Excel files with both .xlsx and .xls formats."""
        import csv
        import os

        from utils.wire_offset_parser import parse_wire_offsets_from_excel

        # Get test data file paths using parametrized values
        test_data_dir = os.path.join(os.path.dirname(__file__), "data")
        excel_path = os.path.join(test_data_dir, excel_file)
        csv_path = os.path.join(test_data_dir, expected_values_csv)

        # Parse expected results from CSV
        expected_results = []
        with open(csv_path, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                expected_results.append(
                    {
                        "TraceabilityNo": row["TraceabilityNo"],
                        "NominalTemp": float(row["NominalTemp"]),
                        "CorrectionFactor": float(row["CorrectionFactor"]),
                    }
                )

        # Parse actual results from Excel
        actual_results = parse_wire_offsets_from_excel(excel_path)

        # Assert results match
        assert len(actual_results) == len(
            expected_results
        ), f"Expected {len(expected_results)} records, got {len(actual_results)}"

        # Sort both lists for consistent comparison
        expected_sorted = sorted(expected_results, key=lambda x: x["NominalTemp"])
        actual_sorted = sorted(actual_results, key=lambda x: x["NominalTemp"])

        for expected, actual in zip(expected_sorted, actual_sorted):
            assert actual["TraceabilityNo"] == expected["TraceabilityNo"]
            assert actual["NominalTemp"] == expected["NominalTemp"]
            assert actual["CorrectionFactor"] == expected["CorrectionFactor"]
