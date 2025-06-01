# utils/sharepoint_client.py
"""SharePoint client for Microsoft Graph API integration."""

import os
import requests
from typing import Optional, Dict, List, Any
import io

# For testing purposes, allow mocking of request extraction
try:
    from flask import request, has_request_context

    _flask_available = True
except ImportError:
    _flask_available = False
    request = None
    has_request_context = None

# Import our graph auth utility
from .graph_auth import get_app_only_token


class SharePointClient:
    """Client for accessing SharePoint files via Microsoft Graph API."""

    def __init__(self, access_token: Optional[str] = None):
        """Initialize SharePoint client with access token.

        Args:
            access_token: Bearer token for Microsoft Graph API authentication.
                         If not provided, will automatically get app-only token.
                         Falls back to Flask request token as last resort.

        Raises:
            ValueError: If no access token provided and none can be obtained.
            RuntimeError: If app-only token acquisition fails.
        """
        if access_token:
            self.access_token = access_token
        else:
            # Try to get app-only token first (preferred for headless operation)
            try:
                self.access_token = get_app_only_token()
            except (ValueError, RuntimeError) as e:
                # Fall back to Flask request token if app token fails
                if (
                    _flask_available
                    and has_request_context
                    and has_request_context()
                    and request
                ):
                    auth_header = request.headers.get("Authorization", "")
                    if auth_header.startswith("Bearer "):
                        self.access_token = auth_header.replace("Bearer ", "")
                    else:
                        raise ValueError(
                            f"Failed to get app token ({e}) and no valid request token found"
                        )
                else:
                    raise ValueError(
                        f"Failed to get app token ({e}) and no Flask request context available"
                    )

        self.base_url = "https://graph.microsoft.com/v1.0"

        # Load SharePoint configuration from environment
        self.site_id = os.environ.get(
            "SHAREPOINT_SITE_ID",
            "jgiquality.sharepoint.com,b8d7ad55-622f-41e1-9140-35b87b4616f9,160cda33-41a0-4b31-8ebf-11196986b3e3",
        )
        self.drive_id = os.environ.get(
            "SHAREPOINT_DRIVE_ID",
            "b!34PQK-JF0EmH57ieExSqveCp2B5j30NMsNTGcMEXae_5x8SnfJhdR6JqUh5dD03F",
        )
        self.pyro_standards_drive_id = os.environ.get(
            "SHAREPOINT_PYRO_STANDARDS_DRIVE_ID"
        )

    def _make_request(self, method: str, endpoint: str, **kwargs) -> requests.Response:
        """Make authenticated request to Microsoft Graph API.

        Args:
            method: HTTP method (GET, POST, etc.)
            endpoint: API endpoint relative to base URL
            **kwargs: Additional request parameters

        Returns:
            Response object from requests library
        """
        url = f"{self.base_url}/{endpoint}"
        headers = kwargs.pop("headers", {})
        headers["Authorization"] = f"Bearer {self.access_token}"

        return requests.request(method, url, headers=headers, **kwargs)

    def get_site_info(self) -> Dict[str, Any]:
        """Get SharePoint site information.

        Returns:
            Dictionary containing site metadata
        """
        response = self._make_request("GET", f"sites/{self.site_id}")
        response.raise_for_status()
        return response.json()

    def get_drive_items(
        self, drive_id: str, folder_path: str = ""
    ) -> List[Dict[str, Any]]:
        """Get items in a SharePoint drive folder.

        Args:
            drive_id: SharePoint drive identifier
            folder_path: Path to folder within drive (empty for root)

        Returns:
            List of drive items
        """
        if folder_path:
            endpoint = f"drives/{drive_id}/root:/{folder_path}:/children"
        else:
            endpoint = f"drives/{drive_id}/root/children"

        response = self._make_request("GET", endpoint)
        response.raise_for_status()
        data = response.json()
        return data.get("value", [])

    def get_file_by_path(self, file_path: str, drive_id: str) -> Dict[str, Any]:
        """Get file metadata by path.

        Args:
            file_path: Path to file within drive
            drive_id: SharePoint drive identifier

        Returns:
            File metadata dictionary
        """
        endpoint = f"drives/{drive_id}/root:/{file_path}"
        response = self._make_request("GET", endpoint)
        if response.status_code == 404:
            return {}
        response.raise_for_status()
        return response.json()

    def get_file_by_id(self, file_id: str, drive_id: str) -> Dict[str, Any]:
        """Get file metadata by ID.

        Args:
            file_id: SharePoint file identifier
            drive_id: SharePoint drive identifier

        Returns:
            File metadata dictionary
        """
        endpoint = f"drives/{drive_id}/items/{file_id}"
        response = self._make_request("GET", endpoint)
        response.raise_for_status()
        return response.json()

    def search_files(self, query: str, drive_id: str) -> List[Dict[str, Any]]:
        """Search for files in a SharePoint drive.

        Args:
            query: Search query string
            drive_id: SharePoint drive identifier

        Returns:
            List of matching files
        """
        endpoint = f"drives/{drive_id}/root/search(q='{query}')"
        response = self._make_request("GET", endpoint)
        response.raise_for_status()
        data = response.json()
        return data.get("value", [])

    def get_file_download_url(self, file_id: str, drive_id: str) -> str:
        """Get temporary download URL for a file.

        Args:
            file_id: SharePoint file identifier
            drive_id: SharePoint drive identifier

        Returns:
            Temporary download URL
        """
        endpoint = f"drives/{drive_id}/items/{file_id}/content"
        response = self._make_request("GET", endpoint, allow_redirects=False)
        return response.headers.get("Location", "")

    def get_file_reference(
        self, file_identifier: str, drive_id: str, by_path: bool = True
    ) -> Dict[str, Any]:
        """Get comprehensive file reference with metadata and download URL.

        Args:
            file_identifier: File path or ID
            drive_id: SharePoint drive identifier
            by_path: Whether file_identifier is a path (True) or ID (False)

        Returns:
            Dictionary with file metadata, download URL, and reference info
        """
        # Get file metadata
        if by_path:
            file_info = self.get_file_by_path(file_identifier, drive_id)
        else:
            file_info = self.get_file_by_id(file_identifier, drive_id)

        # Get download URL
        file_id = file_info.get("id")
        download_url = (
            self.get_file_download_url(file_id, drive_id) if file_id else None
        )  # Build comprehensive reference
        return {
            "id": file_info.get("id"),
            "name": file_info.get("name"),
            "webUrl": file_info.get("webUrl"),
            "downloadUrl": download_url,
            "size": file_info.get("size"),
            "lastModifiedDateTime": file_info.get("lastModifiedDateTime"),
            "createdDateTime": file_info.get("createdDateTime"),
            "mimeType": file_info.get("file", {}).get("mimeType"),
            "driveId": drive_id,
            "path": file_identifier if by_path else None,
        }

    def download_file_content(self, file_id: str, drive_id: str) -> bytes:
        """Download file content as bytes.

        Args:
            file_id: SharePoint file identifier
            drive_id: SharePoint drive identifier

        Returns:
            File content as bytes
        """
        download_url = self.get_file_download_url(file_id, drive_id)
        if not download_url:
            raise ValueError(f"No download URL available for file {file_id}")

        response = requests.get(download_url)
        response.raise_for_status()
        return response.content

    def get_wiresetcerts_file(self) -> Dict[str, Any]:
        """Get WireSetCerts.xlsx file reference from Pyro drive.

        This is an internal method that uses the hardcoded drive ID.

        Returns:
            Dictionary with file metadata and download info

        Raises:
            ValueError: If file not found or drive not configured
        """
        if not self.drive_id:
            raise ValueError(
                "SHAREPOINT_DRIVE_ID not configured"
            )  # Try to get the file by direct path in Shared Documents/Pyro folder
        file_path = "Shared Documents/Pyro/WireSetCerts.xlsx"
        file_ref = self.get_file_reference(file_path, self.drive_id, by_path=True)

        # If file not found by path, try searching for it
        if not file_ref.get("id"):
            search_results = self.search_files("WireSetCerts.xlsx", self.drive_id)
            if search_results:
                # Use the first matching file
                file_info = search_results[0]
                file_id = file_info.get("id")
                if file_id:
                    return self.get_file_reference(
                        file_id, self.drive_id, by_path=False
                    )

            # If still not found, raise error
            raise ValueError("WireSetCerts.xlsx file not found in SharePoint")

        return file_ref


# Convenience functions for external use


def get_wiresetcerts_file_reference(
    access_token: Optional[str] = None,
) -> Dict[str, Any]:
    """Get WireSetCerts.xlsx file reference from SharePoint.

    Args:
        access_token: Bearer token for authentication (optional with new auth)

    Returns:
        File reference dictionary
    """
    client = SharePointClient(access_token)
    return client.get_wiresetcerts_file()


def get_pyro_file_reference(file_path: str, access_token: str) -> Dict[str, Any]:
    """Get file reference from Pyro drive.

    Args:
        file_path: Path to file within Pyro drive
        access_token: Bearer token for authentication

    Returns:
        File reference dictionary
    """
    client = SharePointClient(access_token)
    if not client.drive_id:
        raise ValueError("SHAREPOINT_DRIVE_ID not configured")

    return client.get_file_reference(file_path, client.drive_id, by_path=True)


def get_pyro_standards_file_reference(
    file_path: str, access_token: str
) -> Dict[str, Any]:
    """Get file reference from Pyro Standards drive.

    Args:
        file_path: Path to file within Pyro Standards drive
        access_token: Bearer token for authentication

    Returns:
        File reference dictionary
    """
    client = SharePointClient(access_token)
    if not client.pyro_standards_drive_id:
        raise ValueError("SHAREPOINT_PYRO_STANDARDS_DRIVE_ID not configured")

    return client.get_file_reference(
        file_path, client.pyro_standards_drive_id, by_path=True
    )


def search_pyro_files(query: str, access_token: str) -> List[Dict[str, Any]]:
    """Search for files in Pyro drive.

    Args:
        query: Search query string
        access_token: Bearer token for authentication

    Returns:
        List of matching files
    """
    client = SharePointClient(access_token)
    if not client.drive_id:
        raise ValueError("SHAREPOINT_DRIVE_ID not configured")

    return client.search_files(query, client.drive_id)


def list_pyro_folder_contents(
    folder_path: str, access_token: str
) -> List[Dict[str, Any]]:
    """List contents of a folder in Pyro drive.

    Args:
        folder_path: Path to folder within Pyro drive
        access_token: Bearer token for authentication

    Returns:
        List of folder contents
    """
    client = SharePointClient(access_token)
    if not client.drive_id:
        raise ValueError("SHAREPOINT_DRIVE_ID not configured")

    return client.get_drive_items(client.drive_id, folder_path)


def get_wiresetcerts_content(access_token: Optional[str] = None) -> Dict[str, Any]:
    """Get WireSetCerts.xlsx content parsed as Excel data.

    Args:
        access_token: Bearer token for authentication (optional with new auth)

    Returns:
        Dictionary containing Excel file data:
        - sheet_names: List of sheet names
        - total_sheets: Number of sheets
        - sheets: Dict mapping sheet names to their data
        - file_info: File metadata

    Raises:
        ValueError: If file not found or cannot be accessed
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel file parsing")

    # Get file reference
    client = SharePointClient(access_token)
    file_ref = client.get_wiresetcerts_file()

    if not file_ref.get("id"):
        raise ValueError("WireSetCerts.xlsx file not found in SharePoint")

    # Download file content
    file_content = client.download_file_content(file_ref["id"], client.drive_id)

    # Parse Excel file
    workbook = load_workbook(io.BytesIO(file_content), read_only=True)

    result = {
        "sheet_names": workbook.sheetnames,
        "total_sheets": len(workbook.sheetnames),
        "sheets": {},
        "file_info": {
            "name": file_ref.get("name"),
            "size": file_ref.get("size"),
            "last_modified": file_ref.get("lastModifiedDateTime"),
        },
    }

    # Extract data from each sheet (first few rows as sample)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get headers (first row)
        headers = []
        sample_data = []

        if sheet.max_row and sheet.max_row > 0:
            try:
                # Try to get the first row for headers
                first_row = list(
                    sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                )
                if first_row and first_row[0]:
                    headers = list(first_row[0])

                # Get sample data (first 5 rows after header)
                for row_num in range(2, min(sheet.max_row + 1, 7)):  # Rows 2-6
                    try:
                        row = list(
                            sheet.iter_rows(
                                min_row=row_num, max_row=row_num, values_only=True
                            )
                        )
                        if row and row[0]:
                            sample_data.append(list(row[0]))
                    except (IndexError, TypeError):
                        # Skip rows that can't be accessed
                        continue
            except (IndexError, TypeError):
                # If we can't access any rows, leave headers and sample_data as empty lists
                pass

        result["sheets"][sheet_name] = {
            "headers": headers,
            "sample_data": sample_data,
            "total_rows": sheet.max_row or 0,
            "total_columns": sheet.max_column or 0,
        }

    workbook.close()
    return result


def get_pyro_standards_excel_file(
    file_name: str, access_token: Optional[str] = None
) -> Dict[str, Any]:
    """Get Excel file content from Pyro_Standards folder.

    Args:
        file_name: Name of the Excel file (e.g., "K6_0824.xlsm")
        access_token: Bearer token for authentication (optional with app auth)

    Returns:
        Dictionary containing Excel file data:
        - sheet_names: List of sheet names
        - total_sheets: Number of sheets
        - sheets: Dict mapping sheet names to their data
        - file_info: File metadata

    Raises:
        ValueError: If file not found or cannot be accessed
        ImportError: If openpyxl is not available
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for Excel file parsing")

    client = SharePointClient(
        access_token
    )  # First try direct path in Pyro_Standards folder
    file_path = f"Pyro/Pyro_Standards/{file_name}"

    try:
        file_ref = client.get_file_reference(file_path, client.drive_id, by_path=True)
        if file_ref.get("id"):
            # Direct path worked, we're done
            pass
        else:
            raise ValueError("Direct path failed")
    except:
        # If direct path fails, try listing the Pyro_Standards folder contents
        try:
            folder_contents = client.get_drive_items(
                client.drive_id, "Pyro/Pyro_Standards"
            )
            matching_files = [
                f
                for f in folder_contents
                if f.get("name", "").lower() == file_name.lower()
            ]

            if matching_files:
                file_info = matching_files[0]
                file_id = file_info.get("id")
                if not file_id:
                    raise ValueError("File ID not found")
                file_ref = client.get_file_reference(
                    file_id, client.drive_id, by_path=False
                )
            else:
                # Last resort: search across entire drive (without folder filtering)
                search_results = client.search_files(file_name, client.drive_id)
                if not search_results:
                    raise ValueError(f"{file_name} file not found in SharePoint")

                # Use first exact name match or first result
                exact_matches = [
                    f
                    for f in search_results
                    if f.get("name", "").lower() == file_name.lower()
                ]
                file_info = exact_matches[0] if exact_matches else search_results[0]

                file_id = file_info.get("id")
                if not file_id:
                    raise ValueError(f"{file_name} file not found in SharePoint")

                file_ref = client.get_file_reference(
                    file_id, client.drive_id, by_path=False
                )
        except Exception as e:
            raise ValueError(f"{file_name} file not found in SharePoint: {str(e)}")

    if not file_ref.get("id"):
        raise ValueError(f"{file_name} file not found in SharePoint")

    # Download file content
    file_content = client.download_file_content(file_ref["id"], client.drive_id)

    # Parse Excel file
    workbook = load_workbook(io.BytesIO(file_content), read_only=True)

    result = {
        "sheet_names": workbook.sheetnames,
        "total_sheets": len(workbook.sheetnames),
        "sheets": {},
        "file_info": {
            "name": file_ref.get("name"),
            "size": file_ref.get("size"),
            "last_modified": file_ref.get("lastModifiedDateTime"),
            "path": file_ref.get("path"),
            "web_url": file_ref.get("webUrl"),
        },
    }

    # Extract data from each sheet (first few rows as sample)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Get headers (first row)
        headers = []
        sample_data = []

        if sheet.max_row and sheet.max_row > 0:
            try:
                # Try to get the first row for headers
                first_row = list(
                    sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                )
                if first_row and first_row[0]:
                    headers = list(first_row[0])

                # Get sample data (first 5 rows after header)
                for row_num in range(2, min(sheet.max_row + 1, 7)):  # Rows 2-6
                    try:
                        row = list(
                            sheet.iter_rows(
                                min_row=row_num, max_row=row_num, values_only=True
                            )
                        )
                        if row and row[0]:
                            sample_data.append(list(row[0]))
                    except (IndexError, TypeError):
                        # Skip rows that can't be accessed
                        continue
            except (IndexError, TypeError):
                # If we can't access any rows, leave headers and sample_data as empty lists
                pass

        result["sheets"][sheet_name] = {
            "headers": headers,
            "sample_data": sample_data,
            "total_rows": sheet.max_row or 0,
            "total_columns": sheet.max_column or 0,
        }

    workbook.close()
    return result


def list_pyro_standards_excel_files(
    access_token: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """List all Excel files in the Pyro_Standards folder.

    Args:
        access_token: Bearer token for authentication (optional with app-only auth)

    Returns:
        List of Excel files with metadata

    Raises:
        ValueError: If folder not accessible
    """
    client = SharePointClient(access_token)

    # Search for common Excel file terms
    search_terms = ["xlsx", "excel", "calibration", "standards", "pyro"]
    all_files = []

    for term in search_terms:
        try:
            search_results = client.search_files(term, client.drive_id)
            all_files.extend(search_results)
        except Exception:
            continue  # Skip failed searches
    # Remove duplicates based on file ID and filter for Excel files in Pyro_Standards
    seen_ids = set()
    pyro_standards_files = []

    for result in all_files:
        file_id = result.get("id")
        if file_id in seen_ids:
            continue
        seen_ids.add(file_id)

        file_name = result.get("name", "")

        # Check if it's an Excel file
        is_excel = file_name.lower().endswith((".xlsx", ".xls", ".xlsm"))
        if not is_excel:
            continue

        # Get detailed file info to check parent folder
        try:
            detailed_info = client.get_file_by_id(file_id, client.drive_id)
            parent_ref = detailed_info.get("parentReference", {})
            parent_path = parent_ref.get("path", "")
            parent_name = parent_ref.get("name", "")

            # Check if it's in Pyro_Standards folder
            is_in_pyro_standards = (
                "pyro_standards" in parent_path.lower()
                or parent_name.lower() == "pyro_standards"
            )

            if is_in_pyro_standards:
                pyro_standards_files.append(
                    {
                        "name": result.get("name"),
                        "id": result.get("id"),
                        "size": result.get("size"),
                        "last_modified": result.get("lastModifiedDateTime"),
                        "web_url": result.get("webUrl"),
                        "path": parent_path,
                    }
                )
        except Exception:
            # If we can't get detailed info, skip this file
            continue

    return pyro_standards_files
