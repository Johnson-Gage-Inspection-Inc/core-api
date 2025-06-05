# utils/wire_set_cert_refresher.py
"""
Utility module for refreshing wire set certificate data from SharePoint.

TODO: This module should handle:
1. Downloading WireSetCerts.xlsx from SharePoint Pyro folder
2. Parsing the Excel file to extract serial number mappings
3. Updating the wire_set_certs database table
4. Providing refresh status and error reporting
"""

import logging
from typing import Dict, List, Optional

from sqlalchemy.orm import Session

from db.models import WireSetCert
from utils.database import SessionLocal
from utils.sharepoint_client import SharePointClient


class WireSetCertRefresher:
    """Handles refreshing wire set certificate data from SharePoint."""

    def __init__(self, session: Optional[Session] = None):
        """
        Initialize the refresher.

        Args:
            session: Optional database session. If None, creates a new one.
        """
        self.session = session or SessionLocal()
        self.logger = logging.getLogger(__name__)

        # TODO: Configure these based on your SharePoint structure
        self.sharepoint_site = None  # Will be set from config
        self.pyro_folder_path = "Shared Documents/Pyro"  # TODO: Verify this path
        self.wiresetcerts_filename = "WireSetCerts.xlsx"

    def refresh_wire_set_certs(self) -> Dict:
        """
        Refresh wire set certificate data from SharePoint.

        Returns:
            Dict containing refresh operation results
        """
        result = {
            "status": "success",
            "message": "",
            "records_processed": 0,
            "records_added": 0,
            "records_updated": 0,
            "errors": [],
        }

        try:
            # Download the Excel file from SharePoint
            excel_data = self._download_wiresetcerts_file()

            if not excel_data:
                result["status"] = "error"
                result["message"] = "Failed to download WireSetCerts.xlsx"
                return result

            # Parse the Excel file
            wire_set_mappings = self._parse_wiresetcerts_excel(excel_data)

            if not wire_set_mappings:
                result["status"] = "warning"
                result["message"] = "No wire set mappings found in file"
                return result

            # Update the database
            update_result = self._update_wire_set_certs_table(wire_set_mappings)
            result.update(update_result)

            self.logger.info(f"Wire set cert refresh completed: {result}")

        except Exception as e:
            self.logger.error(f"Error during wire set cert refresh: {e}")
            result["status"] = "error"
            result["message"] = f"Refresh failed: {str(e)}"
            result["errors"].append(str(e))

        return result

    def _download_wiresetcerts_file(self) -> Optional[bytes]:
        """
        Download the WireSetCerts.xlsx file from SharePoint.

        Returns:
            File content as bytes, or None if download failed
        """
        try:
            # Use the existing SharePoint client to get file content
            sharepoint_client = SharePointClient()

            # Get the file reference first
            file_ref = sharepoint_client.get_wiresetcerts_file()

            if not file_ref.get("id"):
                self.logger.error("WireSetCerts.xlsx file not found in SharePoint")
                return None

            # Download the file content as bytes
            file_content = sharepoint_client.download_file_content(
                file_ref["id"], sharepoint_client.drive_id
            )

            self.logger.info(
                f"Successfully downloaded WireSetCerts.xlsx ({len(file_content)} bytes)"
            )
            return file_content

        except Exception as e:
            self.logger.error(f"Error downloading WireSetCerts.xlsx: {e}")
            return None

    def _parse_wiresetcerts_excel(self, excel_data: bytes) -> List[Dict]:
        """
        Parse the WireSetCerts.xlsx file to extract serial number mappings.

        Args:
            excel_data: Raw Excel file content

        Returns:
            List of dictionaries with serial_number and wire_set_group mappings
        """
        try:
            import io

            from openpyxl import load_workbook

            # Load the Excel file from bytes
            workbook = load_workbook(io.BytesIO(excel_data), read_only=True)

            self.logger.info(
                f"Successfully loaded Excel file with sheets: {workbook.sheetnames}"
            )

            mappings = []

            # Process each sheet in the workbook
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self.logger.info(
                    f"Processing sheet '{sheet_name}' with {sheet.max_row} rows"
                )

                if not sheet.max_row or sheet.max_row < 2:
                    self.logger.warning(f"Sheet '{sheet_name}' has no data rows")
                    continue

                # Get headers from first row
                headers_row = list(
                    sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                )[0]
                headers = [
                    str(cell).strip().lower() if cell else "" for cell in headers_row
                ]

                self.logger.info(
                    f"Sheet headers: {headers}"
                )  # Find columns for all required fields
                column_mapping = {}

                # Define column keywords for each field
                column_keywords = {
                    "serial_number": [
                        "serial",
                        "serial_number",
                        "serial number",
                        "wire_serial",
                        "wire serial",
                    ],
                    "wire_set_group": [
                        "type",
                        "group",
                        "wire_set_group",
                        "wire set group",
                        "wire_set",
                        "wire set",
                    ],
                    "asset_id": ["asset_id", "asset id", "assetid", "id"],
                    "asset_tag": ["asset_tag", "asset tag", "assettag", "tag"],
                    "custom_order_number": [
                        "custom_order_number",
                        "custom order number",
                        "order_number",
                        "order number",
                        "custom_order",
                        "custom order",
                    ],
                    "service_date": [
                        "service_date",
                        "service date",
                        "servicedate",
                        "calibration_date",
                        "calibration date",
                    ],
                    "next_service_date": [
                        "next_service_date",
                        "next service date",
                        "nextservicedate",
                        "next_calibration",
                        "next calibration",
                    ],
                    "certificate_number": [
                        "certificate_number",
                        "certificate number",
                        "cert_number",
                        "cert number",
                        "certificate",
                    ],
                    "wire_roll_cert_number": [
                        "wire_roll_cert_number",
                        "wire roll cert number",
                        "roll_cert",
                        "roll cert",
                        "wire_roll_certificate",
                    ],
                }

                # Find column indices for each field
                for field, keywords in column_keywords.items():
                    for i, header in enumerate(headers):
                        if any(keyword in header for keyword in keywords):
                            column_mapping[field] = i
                            break

                self.logger.info(f"Column mapping found: {column_mapping}")

                # Check if we have the required columns (serial_number and wire_set_group)
                if (
                    "serial_number" not in column_mapping
                    or "wire_set_group" not in column_mapping
                ):
                    self.logger.warning(
                        f"Could not find required columns in sheet '{sheet_name}'. "
                        f"Found columns: {list(column_mapping.keys())}"
                    )
                    continue  # Process data rows
                for row_num in range(2, sheet.max_row + 1):
                    try:
                        row_data = list(
                            sheet.iter_rows(
                                min_row=row_num, max_row=row_num, values_only=True
                            )
                        )[0]

                        # Extract all fields using column mapping
                        record = {}

                        for field, col_index in column_mapping.items():
                            if col_index < len(row_data):
                                value = row_data[col_index]

                                # Handle different data types
                                if field in ["service_date", "next_service_date"]:
                                    # Handle datetime fields
                                    if value:
                                        from datetime import datetime

                                        if isinstance(value, datetime):
                                            record[field] = value
                                        elif isinstance(value, str):
                                            try:
                                                # Try to parse string dates
                                                record[field] = datetime.strptime(
                                                    value.strip(), "%Y-%m-%d"
                                                )
                                            except ValueError:
                                                try:
                                                    record[field] = datetime.strptime(
                                                        value.strip(), "%m/%d/%Y"
                                                    )
                                                except ValueError:
                                                    record[field] = None
                                        else:
                                            record[field] = None
                                    else:
                                        record[field] = None

                                elif field == "asset_id":
                                    # Handle integer field
                                    if value:
                                        try:
                                            record[field] = (
                                                int(float(str(value).strip()))
                                                if str(value).strip()
                                                else None
                                            )
                                        except (ValueError, TypeError):
                                            record[field] = None
                                    else:
                                        record[field] = None

                                else:
                                    # Handle string fields
                                    if value:
                                        record[field] = str(value).strip()
                                    else:
                                        record[field] = None

                        # Skip rows without required fields
                        if not record.get("serial_number") or not record.get(
                            "wire_set_group"
                        ):
                            continue

                        # Clean up required string values
                        record["serial_number"] = record["serial_number"].strip()
                        record["wire_set_group"] = record["wire_set_group"].strip()

                        if record["serial_number"] and record["wire_set_group"]:
                            mappings.append(record)

                    except (IndexError, TypeError) as e:
                        self.logger.warning(
                            f"Error processing row {row_num} in sheet '{sheet_name}': {e}"
                        )
                        continue

            workbook.close()

            self.logger.info(
                f"Extracted {len(mappings)} wire set mappings from Excel file"
            )

            # Log a few examples for debugging
            if mappings:
                examples = mappings[:3]
                self.logger.info(f"Example mappings: {examples}")

            return mappings

        except Exception as e:
            self.logger.error(f"Error parsing WireSetCerts.xlsx: {e}")
            return []

    def _update_wire_set_certs_table(self, mappings: List[Dict]) -> Dict:
        """
        Update the wire_set_certs table with new mappings.

        Args:
            mappings: List of serial number to wire set group mappings

        Returns:
            Dict with update statistics
        """
        result = {
            "records_processed": 0,
            "records_added": 0,
            "records_updated": 0,
            "errors": [],
        }

        try:
            # Start a transaction
            for mapping in mappings:
                result["records_processed"] += 1  # Check if record already exists
                existing = (
                    self.session.query(WireSetCert)
                    .filter_by(serial_number=mapping["serial_number"])
                    .first()
                )

                if existing:
                    # Update existing record if any field has changed
                    updated = False

                    # Check and update all fields
                    fields_to_check = [
                        "wire_set_group",
                        "asset_id",
                        "asset_tag",
                        "custom_order_number",
                        "service_date",
                        "next_service_date",
                        "certificate_number",
                        "wire_roll_cert_number",
                    ]

                    for field in fields_to_check:
                        new_value = mapping.get(field)
                        current_value = getattr(existing, field, None)

                        if new_value != current_value:
                            setattr(existing, field, new_value)
                            updated = True
                            self.logger.debug(
                                f"Updated {field} for serial {mapping['serial_number']}: "
                                f"{current_value} -> {new_value}"
                            )

                    if updated:
                        result["records_updated"] += 1
                        self.logger.info(f"Updated serial {mapping['serial_number']}")

                else:
                    # Insert new record with all fields
                    new_cert = WireSetCert(
                        serial_number=mapping["serial_number"],
                        wire_set_group=mapping["wire_set_group"],
                        asset_id=mapping.get("asset_id"),
                        asset_tag=mapping.get("asset_tag"),
                        custom_order_number=mapping.get("custom_order_number"),
                        service_date=mapping.get("service_date"),
                        next_service_date=mapping.get("next_service_date"),
                        certificate_number=mapping.get("certificate_number"),
                        wire_roll_cert_number=mapping.get("wire_roll_cert_number"),
                    )
                    self.session.add(new_cert)
                    result["records_added"] += 1
                    self.logger.info(
                        f"Added new serial {mapping['serial_number']}: {mapping['wire_set_group']}"
                    )

            # Commit the transaction
            self.session.commit()

            self.logger.info(
                f"Database update completed: "
                f"{result['records_added']} added, "
                f"{result['records_updated']} updated, "
                f"{result['records_processed']} processed"
            )

        except Exception as e:
            # Rollback on error
            self.session.rollback()
            self.logger.error(f"Error updating wire_set_certs table: {e}")
            result["errors"].append(str(e))

        return result


def refresh_wire_set_certs() -> Dict:
    """
    Convenience function to refresh wire set certificates.

    TODO: This is the main entry point for the refresh operation.
    Called by the API endpoint to trigger the refresh.

    Returns:
        Dict containing refresh operation results
    """
    refresher = WireSetCertRefresher()
    try:
        return refresher.refresh_wire_set_certs()
    finally:
        refresher.session.close()


# TODO: Consider adding these additional functions:
# - validate_wire_set_mappings(mappings: List[Dict]) -> List[str]
# - get_wire_set_for_serial(serial_number: str) -> Optional[str]
# - export_wire_set_certs_to_csv() -> str
# - get_wire_set_for_serial(serial_number: str) -> Optional[str]
# - export_wire_set_certs_to_csv() -> str
