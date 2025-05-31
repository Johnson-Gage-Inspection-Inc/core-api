#!/usr/bin/env python3
# utils/excel_parser.py
"""
Excel file parsing utilities for daqbook offset data.
Handles parsing of calibration data from Excel (.xlsm) files.
"""

from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from pathlib import Path

def parse_daqbook_offsets_from_excel(filepath: str, tn: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Parse daqbook offset data from an Excel file using Power Query logic.
    
    Power Query Logic:
    - Temperatures are in A42:A47 (6 temperatures)
    - Block offsets are at rows: {42, 50, 60, 68, 78, 86, 96} (7 blocks)
    - Channels are in columns B-G (6 channels = points 1-6)
    - Each temperature appears in each block, at block_offset + temp_index
    
    Args:
        filepath: Path to the Excel file
        tn: Test number/identifier (if None, will try to extract from filename)
        
    Returns:
        List of offset data dictionaries with fields: Temp, Point, Reading, Delta
    """
    if tn is None:
        # Try to extract TN from filename
        filename = Path(filepath).stem
        tn = filename.replace("_", "").replace("-", "")

    try:
        workbook = load_workbook(filepath, read_only=True, data_only=True)
        all_offsets = []
        
        # Use Sheet1 as specified in the Power Query logic
        sheet = None
        if "Sheet1" in workbook.sheetnames:
            sheet = workbook["Sheet1"]
        elif workbook.sheetnames:
            sheet = workbook[workbook.sheetnames[0]]
        else:
            return []
        
        # Power Query constants
        temp_rows = list(range(42, 48))  # A42:A47 (rows 42-47)
        block_offsets = [42, 50, 60, 68, 78, 86, 96]  # 7 blocks
        channel_columns = list(range(2, 8))  # B-G (columns 2-7)
        
        # Extract temperatures from A42:A47
        temperatures = []
        for row in temp_rows:
            temp_value = sheet.cell(row=row, column=1).value
            if isinstance(temp_value, (int, float)):
                temperatures.append(float(temp_value))
        
        # Parse offset data from each block
        for block_idx, block_start in enumerate(block_offsets):
            for temp_idx, temp in enumerate(temperatures):
                data_row = block_start + temp_idx
                
                for col_idx, column in enumerate(channel_columns):
                    # Calculate point number: (block * 6) + channel + 1
                    point = (block_idx * 6) + col_idx + 1
                    reading_value = sheet.cell(row=data_row, column=column).value
                    if isinstance(reading_value, (int, float)):
                        reading = float(reading_value)
                        # Calculate delta: (Reading - Temp) * -1, rounded to 2 decimal places
                        delta = round((reading - temp) * -1, 2)
                        
                        offset_record = {
                            "Reading": reading,
                            "Temp": temp,
                            "Point": point,
                            "Delta": delta
                        }
                        all_offsets.append(offset_record)
        
        workbook.close()
        return all_offsets
        
    except Exception as e:
        # Return empty list on error - this allows graceful handling
        return []
