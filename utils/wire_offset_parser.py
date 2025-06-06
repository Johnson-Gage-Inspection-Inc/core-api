#!/usr/bin/env python
# utils/wire_offset_parser.py
"""
Wire offset Excel file parsing utilities.
Handles parsing of wire correction factor data from Excel (.xlsx and .xls) files.
"""

import os
import re
from typing import Any, Dict, List

import openpyxl

try:
    import xlrd
except ImportError:
    xlrd = None


def _parse_temperature_cell(cell_value: Any) -> float:
    """
    Given a cell value like '2000.1°F ' (possibly with trailing whitespace),
    strip off the '°F', parentheses if present, and return a float.
    Raises ValueError if the resulting text is not a float.
    """
    if cell_value is None:
        raise ValueError("Temperature cell is empty")
    s = str(cell_value).strip()
    # Remove any parentheses or '°F' markers
    # e.g. "(‐0.1°F)" → "-0.1"
    s = s.replace("(", "").replace(")", "")
    # Remove '°F' or '°f' (case-insensitive)
    s = re.sub(r"°[Ff]", "", s).strip()
    # Now s should look like '2000.1' or '-0.1'
    try:
        return float(s)
    except Exception as e:
        raise ValueError(f"Cannot parse temperature from '{cell_value}'") from e


def parse_wire_offsets_from_excel(path: str) -> List[Dict[str, Any]]:
    """
    Open the given Excel file (.xlsx or .xls) and parse the 'CERT, Wire Roll'
    worksheet to extract wire offset correction factors.

    Structure:
    - Temperatures are in rows 21 and 27, columns C:J (up to 8 temperatures)
    - Correction factors are in rows 24 and 30, columns C:J (corresponding to temperatures)
    - Wire traceability number is extracted from filename

    Returns a list of dicts, each dict:
        {
          "TraceabilityNo": <string extracted from filename>,
          "NominalTemp": <float from rows 21/27>,
          "CorrectionFactor": <float from rows 24/30>
        }
    """
    # Extract traceability number from filename (e.g., "072513A.xlsx" -> "072513A")
    filename = os.path.basename(path)
    traceability_no = os.path.splitext(filename)[0]

    # Check file extension to use the appropriate library
    file_ext = os.path.splitext(path)[1].lower()

    if file_ext == ".xls":
        # Use xlrd for older .xls format
        if xlrd is None:
            raise ImportError(
                "xlrd library is required to parse .xls files. Install with 'pip install xlrd'"
            )

        return _parse_xls_file(path, traceability_no)
    else:
        # Use openpyxl for .xlsx format (default)
        return _parse_xlsx_file(path, traceability_no)


def _parse_xlsx_file(path: str, traceability_no: str) -> List[Dict[str, Any]]:
    """Parse .xlsx file using openpyxl."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "CERT, Wire Roll" not in wb.sheetnames:
        raise ValueError(f"'CERT, Wire Roll' sheet not found in {path}")

    ws = wb["CERT, Wire Roll"]
    results: List[Dict[str, Any]] = []

    # Define temperature and correction factor row/column ranges
    temp_rows = [21, 27]  # Temperatures in rows 21 and 27
    cf_rows = [24, 30]  # Correction factors in rows 24 and 30
    columns = list(range(3, 11))  # Columns C through J (3 through 10)
    # Read temperatures and correction factors from both row sets
    for row_set_idx in range(len(temp_rows)):
        temp_row = temp_rows[row_set_idx]
        cf_row = cf_rows[row_set_idx]

        for col in columns:
            # Read temperature value
            temp_cell = ws.cell(row=temp_row, column=col).value
            if temp_cell is None:
                continue  # Skip empty temperature cells

            # Read corresponding correction factor
            cf_cell = ws.cell(row=cf_row, column=col).value
            if cf_cell is None:
                continue  # Skip if no correction factor

            try:
                # Convert to float values
                nominal_temp = float(temp_cell)
                correction_factor = float(cf_cell)

                results.append(
                    {
                        "TraceabilityNo": traceability_no,
                        "NominalTemp": nominal_temp,
                        "CorrectionFactor": correction_factor,
                    }
                )
            except (ValueError, TypeError):
                # Skip cells that can't be converted to float
                continue

    wb.close()
    return results


def _parse_xls_file(path: str, traceability_no: str) -> List[Dict[str, Any]]:
    """Parse older .xls file format using xlrd."""
    wb = xlrd.open_workbook(path)
    sheet_names = wb.sheet_names()
    if "CERT, Wire Roll" not in sheet_names:
        raise ValueError(f"'CERT, Wire Roll' sheet not found in {path}")

    ws = wb.sheet_by_name("CERT, Wire Roll")
    results: List[Dict[str, Any]] = []

    # Define temperature and correction factor row/column ranges
    temp_rows = [20, 26]  # Temperatures in rows 21 and 27 (0-indexed in xlrd)
    cf_rows = [23, 29]  # Correction factors in rows 24 and 30 (0-indexed in xlrd)
    columns = list(range(2, 10))  # Columns C through J (0-indexed in xlrd)

    # Read temperatures and correction factors from both row sets
    for row_set_idx in range(len(temp_rows)):
        temp_row = temp_rows[row_set_idx]
        cf_row = cf_rows[row_set_idx]

        for col in columns:
            # Read temperature value
            if col >= ws.ncols or temp_row >= ws.nrows:
                continue  # Skip if out of bounds

            temp_cell = ws.cell_value(temp_row, col)
            if temp_cell == "" or temp_cell is None:
                continue  # Skip empty temperature cells

            # Read corresponding correction factor
            if cf_row >= ws.nrows:
                continue  # Skip if out of bounds

            cf_cell = ws.cell_value(cf_row, col)
            if cf_cell == "" or cf_cell is None:
                continue  # Skip if no correction factor

            try:
                # Convert to float values
                nominal_temp = float(temp_cell)
                correction_factor = float(cf_cell)

                results.append(
                    {
                        "TraceabilityNo": traceability_no,
                        "NominalTemp": nominal_temp,
                        "CorrectionFactor": correction_factor,
                    }
                )
            except (ValueError, TypeError):
                # Skip cells that can't be converted to float
                continue

    return results
